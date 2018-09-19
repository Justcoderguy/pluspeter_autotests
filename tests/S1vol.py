#! python3
# Standard checkout - checkout with standard product with one volume, Credit Card

import openpyxl
import random
import os


def init_checkout(login, password):

    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as EC
    from selenium.common.exceptions import NoSuchElementException

    print('Running checkout with standard product with one volume, Credit Card...')
    docs_dir = 'C:\\Users\\pzakharevich\\Desktop\\Image attachments\\Testdata\\'
    pdf_docs = []
    for item in os.listdir(docs_dir):
        if item.endswith('.pdf'):
            pdf_docs.append(item)
        else:
            pass
    for pdf in pdf_docs:
        print(str(pdf_docs.index(pdf)) + ' ' + pdf)
    print('\nType number of the document you want to print.\n')

    # Select document to print:

    select_pdf = True
    while select_pdf:
        try:
            pdf_number = int(input('Standard 1 volume document: '))
            print('Selected pdf ' + pdf_docs[pdf_number])
            select_pdf = False
        except ValueError:
            print('Only numbers are acceptable. Please, try again.')
        except IndexError:
            print('No document with this number. Please, try again.')

    # set volume option - text color:

    while True:
        content_color = str(input('\nSelect print color ("1c" - black and white; "4c" - all colors): '))
        if content_color == '1c' or content_color == '4c':
            break
        else:
            print('Invalid input. Please, try again.')

    # set volume option - binding:

    while True:
        binding = str(input('\nSelect binding ("wireo" - wired; "punched"): '))
        if binding == 'wireo' or binding == 'punched':
            break
        else:
            print('Invalid input. Please, try again.')

    # set volume option - cover color:

    select_cover = True
    while select_cover:
        try:
            print('\n')
            color_options = ['blue', 'green', 'purple', 'red', 'yellow', 'dark_blue', 'light_blue']
            for color in color_options:
                print(str(color_options.index(color)) + ' ' + color)
            color_num = int(input('\nType number of the cover color you want to use: '))
            select_cover = False
        except ValueError:
            print('Invalid input. Please, try again.')
        except IndexError:
            print('Invalid input. Please, try again.')

    # Select payment method:

    while True:
        payment_method = str(input('\nSelect payment method '
                                   '(Enter "pp" for PayPal or "cc" for Credit Card, no quotes): '))
        if payment_method == 'pp' or payment_method == 'cc':
            break
        else:
            print('Invalid input. Please, try again.')

    # Create instance and go to landing page:

    browser = webdriver.Chrome('C:\ChromeDriver\chromedriver.exe')
    browser.get('https://staging.pluspeter.com')
    browser.maximize_window()

    # Go to login page and sign in:

    browser.find_element_by_link_text('Login').click()
    browser.find_element_by_id('id_username').send_keys(login)
    browser.find_element_by_id('id_password').send_keys(password)
    browser.find_element_by_id('id_password').submit()

    # Go to products page and open standard volume:

    WebDriverWait(browser, 7).until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    WebDriverWait(browser, 5).until(EC.element_to_be_clickable((By.LINK_TEXT, 'Jetzt Drucken'))).click()
    WebDriverWait(browser, 7).until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    WebDriverWait(browser, 10).until(EC.visibility_of_element_located((By.XPATH, "//div[@class='button-bottom-holder']"
                                                                                 "/button[@id='print_europe']"))).click()

    # Add document:

    WebDriverWait(browser, 7).until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    WebDriverWait(browser, 5).until(EC.presence_of_element_located((By.NAME, 'file-input')))
    browser.find_element_by_name('file-input').send_keys(docs_dir + pdf_docs[pdf_number])

    # Select and accept document orientation:

    WebDriverWait(browser, 7).until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    orientation_options = browser.find_elements_by_xpath("//span[@class='input-radio']")
    try:
        orientation_options[1].click()
        browser.find_element_by_xpath("//button[@type='submit']").click()
    except IndexError:
        pass

    # Set volume options:

    WebDriverWait(browser, 7).until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    volume_options = browser.find_elements_by_xpath("//div[@class='profil-select select-search-input div-as-input']")

    if content_color == '1c':
        volume_options[0].click()
        browser.find_element_by_xpath("//div[@value='1c']").click()
    else:
        pass

    WebDriverWait(browser, 7).until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))

    if binding == 'punched':
        volume_options[1].click()
        browser.find_element_by_xpath("//div[@value='punched']").click()
    else:
        pass

    WebDriverWait(browser, 7).until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))

    if color_num != color_options.index('blue'):
        volume_options[2].click()
        browser.find_element_by_xpath("//div[@value='" + color_options[color_num] + "']").click()
    else:
        pass

    WebDriverWait(browser, 7).until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    browser.find_element_by_xpath("//input[@placeholder='Titel']").send_keys('Stan/dard by-b0t!')
    browser.find_element_by_id('wizard').click()

    if payment_method == 'cc':

        # Select CC payment method, mark checkboxes and open stripe:

        WebDriverWait(browser, 12).until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
        WebDriverWait(browser, 2).until(EC.presence_of_element_located((By.ID, 'paypal3'))).click()
        browser.find_element_by_xpath("//label[@for='checkbox-2']").click()
        browser.find_element_by_xpath("//label[@for='checkbox-3']").click()
        browser.find_element_by_id('paypalPayment').click()

        # Switch to stripe and finish checkout:

        WebDriverWait(browser, 12).until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
        WebDriverWait(browser, 7).until(EC.frame_to_be_available_and_switch_to_it('stripe_checkout_app'))
        browser.find_element_by_xpath("//input[@placeholder='Card number']").send_keys('4242 4242 4242 4242')
        browser.find_element_by_xpath("//input[@placeholder='MM / YY']").send_keys('0123')
        browser.find_element_by_xpath("//input[@placeholder='CVC']").send_keys('456')
        browser.find_element_by_xpath("//input[@placeholder='ZIP Code']").send_keys('7890')
        browser.find_element_by_xpath("//button[@type='submit']").click()
        browser.switch_to.default_content()
    else:
        # Mark checkboxes and go to Paypal:

        WebDriverWait(browser, 12).until(
            EC.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
        WebDriverWait(browser, 2).until(EC.presence_of_element_located((By.ID, 'paypal3'))).click()
        browser.find_element_by_id('paypal1').click()
        browser.find_element_by_xpath("//label[@for='checkbox-2']").click()
        browser.find_element_by_xpath("//label[@for='checkbox-3']").click()
        browser.find_element_by_id('paypalPayment').click()
        WebDriverWait(browser, 10).until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
        WebDriverWait(browser, 10).until(EC.url_contains('www.sandbox.paypal.com'))

        # PayPal checkout:

        WebDriverWait(browser, 25).until(EC.invisibility_of_element_located((By.ID, 'preloaderSpinner')))
        try:
            browser.find_element_by_xpath("//a[@class='btn full ng-binding']").click()
        except NoSuchElementException:
            print("Log in button not found. Trying to find email field...")
        WebDriverWait(browser, 25).until(EC.invisibility_of_element_located((By.ID, 'preloaderSpinner')))
        WebDriverWait(browser, 10).until(EC.visibility_of_element_located((By.NAME, 'login_email'))).send_keys(
            'info-buyer@printpeter.de')
        WebDriverWait(browser, 10).until(EC.visibility_of_element_located((By.ID, 'btnNext'))).click()
        WebDriverWait(browser, 25).until(EC.invisibility_of_element_located((By.ID, 'preloaderSpinner')))
        WebDriverWait(browser, 10).until(EC.visibility_of_element_located((By.ID, 'password'))).send_keys(
            'fdjkru12')
        browser.find_element_by_id('btnLogin').click()
        WebDriverWait(browser, 30).until(EC.invisibility_of_element_located((By.ID, 'preloaderSpinner')))
        browser.find_element_by_id('confirmButtonTop').click()
        WebDriverWait(browser, 25).until(EC.invisibility_of_element_located((By.ID, 'preloaderSpinner')))

    # Go to profile:

    WebDriverWait(browser, 12).until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    WebDriverWait(browser, 10).until(EC.presence_of_element_located((By.XPATH, "//a[@class='closer']"))).click()
    WebDriverWait(browser, 12).until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))

    # Logout and quit:

    WebDriverWait(browser, 7).until(EC.presence_of_element_located((By.LINK_TEXT, "Abmelden"))).click()
    WebDriverWait(browser, 12).until(EC.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    browser.quit()


if __name__ == "__main__":
    while True:
        creds = str(input('\nEnter credentials flow ("auto" - creds will be selected randomly; '
                          '"man" - type creds manually): '))
        if creds == 'auto':
            print('Auto mode is selected.')
            os.chdir('C:\\Users\\pzakharevich\\Desktop\\Education\\PlusPeter\\docs')
            login_list = []
            password_list = []
            wb = openpyxl.load_workbook('credentials.xlsx')
            wb.save('credentials - Copy.xlsx')
            wb = openpyxl.load_workbook('credentials - Copy.xlsx')
            sheet = wb.active
            if sheet.max_row <= 2:
                login_list.append(sheet.cell(2, 1).value)
                password_list.append(sheet.cell(2, 2).value)
                login = login_list[0]
                password = password_list[0]
            else:
                for row in range(2, sheet.max_row + 1):
                    login_list.append(sheet.cell(row, 1).value)
                    password_list.append(sheet.cell(row, 2).value)
                creds_num = random.randint(0, len(login_list) - 1)
                login = login_list[creds_num]
                password = password_list[creds_num]
            break
        elif creds == 'man':
            print('Manual mode is selected.\n')
            login = input('Login: ')
            password = input('Password: ')
            break
        else:
            print('Invalid input. Please, try again.')

    init_checkout(login, password)
