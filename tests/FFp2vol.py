#! python3
# Freemium and freemium plus checkout
# checkout with freemium and freemium plus or two freemium plus volumes, free

import openpyxl
import random
import os


def init_checkout(user_login, user_password):

    import os
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as ec
    from selenium.common.exceptions import NoSuchElementException, TimeoutException, WebDriverException

    print('Running checkout with freemium and freemium plus product, two volumes, free...')
    test_docs_dir = 'C:\\Users\\pzakharevich\\Desktop\\Image attachments\\Testdata\\'
    pdf_docs = []
    for item in os.listdir(docs_dir):
        if item.endswith('.pdf'):
            pdf_docs.append(item)
        else:
            pass
    for pdf in pdf_docs:
        print(str(pdf_docs.index(pdf)) + ' ' + pdf)
    print('\nType number of the document you want to print.\n')
    while True:
        try:
            pdf_number_1 = int(input('Freemium/Freemium plus 1 volume document: '))
            print('Selected pdf ' + pdf_docs[pdf_number_1])
            pdf_number_2 = int(input('Freemium plus 2 volume document: '))
            print('Selected pdf ' + pdf_docs[pdf_number_2])
            break
        except ValueError:
            print('Only numbers are acceptable. Please, try again.')
        except IndexError:
            print('No document with this number. Please, try again.')

    while True:
        payment_method = str(input('\nSelect payment method '
                                   '(Enter "pp" for PayPal or "cc" for Credit Card, no quotes): '))
        if payment_method == 'pp' or payment_method == 'cc':
            break
        else:
            print('Invalid input. Please, try again.')
            continue

    # Create instance and go to landing page:

    browser = webdriver.Chrome('C:\ChromeDriver\chromedriver.exe')
    browser.get('https://staging.pluspeter.com')
    browser.maximize_window()

    # Go to login page and sign in:

    browser.find_element_by_link_text('Login').click()
    browser.find_element_by_id('id_username').send_keys(user_login)
    browser.find_element_by_id('id_password').send_keys(user_password)
    browser.find_element_by_id('id_password').submit()

    # Go to products page and open freemium volume:

    WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    WebDriverWait(browser, 5).until(ec.element_to_be_clickable((By.LINK_TEXT, 'Jetzt Drucken'))).click()
    WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    WebDriverWait(browser, 10).until(ec.visibility_of_element_located((By.XPATH,
                                                                       "//div[@class='button-bottom-holder']"
                                                                       "/button[@id='print_freemium']"))).click()

    # Add document:

    WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    try:
        browser.find_element_by_id('omnibox-file-input').send_keys(test_docs_dir + pdf_docs[pdf_number_1])
        WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
        browser.find_element_by_xpath('//*[@id="wizard"]/div[2]/div[1]/div/div[4]/input').send_keys('Free/mium by-b0t!')
        browser.find_element_by_xpath('//*[@id="wizard"]').click()
        WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    except NoSuchElementException:
        print('\nFreemium print is not available. Printing freemium plus instead...')
        browser.find_element_by_id('print_freemium_plus').click()
        WebDriverWait(browser, 5).until(ec.presence_of_element_located((By.NAME, 'file-input')))
        browser.find_element_by_name('file-input').send_keys(test_docs_dir + pdf_docs[pdf_number_1])
        WebDriverWait(browser, 12).until(
            ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
        browser.find_element_by_xpath('//*[@id="wizard"]/div[2]/div[2]/'
                                      'div/div/div[2]/input').send_keys('Free/mium by-b0t!')
        browser.find_element_by_xpath('//*[@id="wizard"]').click()

    # Add second freemium plus volume:

    WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    try:
        print('Trying to add second freemium plus volume...')
        WebDriverWait(browser, 5).until(ec.element_to_be_clickable((By.XPATH,
                                                                    "//span[text()="
                                                                    "'Freemium Plus dazu bestellen']"))).click()
    except TimeoutException:
        WebDriverWait(browser, 5).until(ec.element_to_be_clickable((By.XPATH, "//a[@href='#']"))).click()
        WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
        WebDriverWait(browser, 5).until(ec.element_to_be_clickable((By.XPATH, "//div[@class='button-bottom-holder']"
                                                                              "/button[@id='print_freemium']"))).click()
        browser.find_element_by_id('print_freemium_plus').click()

    WebDriverWait(browser, 20).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    browser.execute_script('window.scrollTo(0, 900)')
    file_input = browser.find_elements(By.NAME, 'file-input')
    try:
        file_input[1].send_keys(test_docs_dir + pdf_docs[pdf_number_2])
    except IndexError:
        file_input[0].send_keys(test_docs_dir + pdf_docs[pdf_number_2])
    WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    browser.find_element_by_xpath('//*[@id="wizard"]/div[3]/div[2]/div/div/div[2]/input').send_keys('Free/mium by-b0t!')
    browser.find_element_by_xpath('//*[@id="wizard"]').click()
    WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))

    # Payment:

    if payment_method == 'cc':
        try:
            # Select CC payment method, mark checkboxes and open stripe:

            WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH,
                                                                                 "//div[@class='loading_box']")))
            WebDriverWait(browser, 2).until(ec.presence_of_element_located((By.ID, 'paypal3'))).click()
            browser.find_element_by_xpath("//label[@for='checkbox-1']").click()
            browser.find_element_by_xpath("//label[@for='checkbox-2']").click()
            browser.find_element_by_id('paypalPayment').click()

            # Switch to stripe and finish checkout:

            WebDriverWait(browser, 7).until(ec.frame_to_be_available_and_switch_to_it('stripe_checkout_app'))
            browser.find_element_by_xpath("//input[@placeholder='Card number']").send_keys('4242 4242 4242 4242')
            browser.find_element_by_xpath("//input[@placeholder='MM / YY']").send_keys('0123')
            browser.find_element_by_xpath("//input[@placeholder='CVC']").send_keys('456')
            browser.find_element_by_xpath("//input[@placeholder='ZIP Code']").send_keys('7890')
            browser.find_element_by_xpath("//button[@type='submit']").click()
            browser.switch_to.default_content()

        except TimeoutException:
            print('Product is free. Finishing checkout...')

        # Mark checkboxes and finish checkout:

        try:
            WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH,
                                                                                 "//div[@class='loading_box']")))
            browser.find_element_by_xpath("//label[@for='checkbox-1']").click()
            browser.find_element_by_xpath("//label[@for='checkbox-2']").click()
            browser.find_element_by_id('paypalPayment').click()
        except WebDriverException:
            pass
    else:
        try:
            # Mark checkboxes and go to Paypal:

            WebDriverWait(browser, 12).until(
                ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
            WebDriverWait(browser, 2).until(ec.presence_of_element_located((By.ID, 'paypal3'))).click()
            browser.find_element_by_id('paypal1').click()
            browser.find_element_by_xpath("//label[@for='checkbox-1']").click()
            browser.find_element_by_xpath("//label[@for='checkbox-2']").click()
            browser.find_element_by_id('paypalPayment').click()
            WebDriverWait(browser, 10).until(
                ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
            WebDriverWait(browser, 10).until(ec.url_contains('www.sandbox.paypal.com'))

            # PayPal checkout:

            WebDriverWait(browser, 25).until(ec.invisibility_of_element_located((By.ID, 'preloaderSpinner')))
            try:
                browser.find_element_by_xpath("//a[@class='btn full ng-binding']").click()
            except NoSuchElementException:
                print("Log in button not found. Trying to find email field...")
            WebDriverWait(browser, 25).until(ec.invisibility_of_element_located((By.ID, 'preloaderSpinner')))
            WebDriverWait(browser, 10).until(ec.visibility_of_element_located((By.NAME, 'login_email'))).send_keys(
                'info-buyer@printpeter.de')
            WebDriverWait(browser, 10).until(ec.visibility_of_element_located((By.ID, 'btnNext'))).click()
            WebDriverWait(browser, 25).until(ec.invisibility_of_element_located((By.ID, 'preloaderSpinner')))
            WebDriverWait(browser, 10).until(ec.visibility_of_element_located((By.ID, 'password'))).send_keys(
                'fdjkru12')
            browser.find_element_by_id('btnLogin').click()
            WebDriverWait(browser, 30).until(ec.invisibility_of_element_located((By.ID, 'preloaderSpinner')))
            browser.find_element_by_id('confirmButtonTop').click()
            WebDriverWait(browser, 25).until(ec.invisibility_of_element_located((By.ID, 'preloaderSpinner')))

        except TimeoutException:
            print('Product is free. Finishing checkout...')

            # Mark checkboxes and finish checkout:

        try:
            WebDriverWait(browser, 12).until(
                ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
            browser.find_element_by_xpath("//label[@for='checkbox-1']").click()
            browser.find_element_by_xpath("//label[@for='checkbox-2']").click()
            browser.find_element_by_id('paypalPayment').click()
        except WebDriverException:
            pass

    # Go to profile:

    WebDriverWait(browser, 20).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    WebDriverWait(browser, 20).until(ec.presence_of_element_located((By.XPATH, "//a[@class='closer']"))).click()

    # Logout and quit:

    WebDriverWait(browser, 20).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    WebDriverWait(browser, 7).until(ec.presence_of_element_located((By.LINK_TEXT, "Abmelden"))).click()
    WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    print('\nCheckout with freemium and freemium plus or two freemium plus products is finished successfully.')
    browser.quit()


if __name__ == "__main__":
    docs_dir = 'C:\\Users\\pzakharevich\\Desktop\\Education\\PlusPeter\\docs'
    while True:
        auto_mode = 'Auto'
        manual_mode = 'Manual'
        creds_list = [auto_mode, manual_mode]
        print('\n')
        for mode in creds_list:
            print(str(creds_list.index(mode)) + ' ' + mode)
        while True:
            try:
                mode_num = str(input('\nEnter mode number (Auto will let bot select credentials; Manual will let you'
                                     ' enter credentials): '))
                print('Selected mode ' + creds_list[int(mode_num)])
                break
            except IndexError:
                print('Invalid input. Please, try again.')
            except ValueError:
                print('Only numbers are acceptable. Please, try again.')
        if creds_list[int(mode_num)] == auto_mode:
            country_list = ['Germany', 'Austria', 'Switzerland']
            for country in country_list:
                print(str(country_list.index(country)) + ' ' + country)
            while True:
                try:
                    user_country = str(input('\nEnter user\'s country number: '))
                    print('Selected country ' + country_list[int(user_country)])
                    break
                except IndexError:
                    print('Invalid input. Please, try again.')
            print('Program will select creds automatically.')
            os.chdir(docs_dir)
            login_list = []
            password_list = []
            wb = openpyxl.load_workbook('credentials.xlsx')
            wb.save('credentials - Copy.xlsx')
            wb = openpyxl.load_workbook('credentials - Copy.xlsx')
            if country_list[int(user_country)] == 'Germany':
                sheet = wb['German users']
            elif country_list[int(user_country)] == 'Austria':
                sheet = wb['Austrian users']
            else:
                sheet = wb['Swiss users']
            if sheet.max_row <= 2:
                login_list.append(sheet.cell(2, 1).value)
                password_list.append(sheet.cell(2, 2).value)
                login = login_list[0]
                password = password_list[0]
            else:
                for row in range(2, sheet.max_row + 1):
                    login_list.append(sheet.cell(row, 1).value)
                    password_list.append(sheet.cell(row, 2).value)
                creds_num = random.randint(0, len(login_list) - 1)
                login = login_list[creds_num]
                password = password_list[creds_num]
            break
        elif creds_list[int(mode_num)] == manual_mode:
            print('Manual mode is selected.\n')
            login = input('Login: ')
            password = input('Password: ')
            break
        else:
            print('Invalid input. Please, try again.')

    init_checkout(login, password)
