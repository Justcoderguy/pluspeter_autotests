#! python3
# Freemium checkout - checkout with freemium product, one volume

import openpyxl
import random
import os
import logging

docs_dir = 'C:\\Users\\pzakharevich\\Desktop\\Education\\PlusPeter\\docs'

logging.basicConfig(level=logging.DEBUG, filename='FFp1vol.log', filemode='w',
                    format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%d/%m/%Y %I:%M:%S')

# logging.disable(logging.DEBUG)


def init_checkout(user_login, user_password):

    import os
    from selenium import webdriver
    from selenium.webdriver.common.by import By
    from selenium.webdriver.support.ui import WebDriverWait
    from selenium.webdriver.support import expected_conditions as ec
    from selenium.common.exceptions import NoSuchElementException, WebDriverException, TimeoutException

    print('Running checkout with freemium or freemium plus product.')
    test_docs_dir = 'C:\\Users\\pzakharevich\\Desktop\\Image attachments\\Testdata\\'
    paypal_login = 'info-buyer@printpeter.de'
    paypal_password = 'fdjkru12'
    pdf_docs = []
    for item in os.listdir(test_docs_dir):
        if item.endswith('.pdf'):
            pdf_docs.append(item)
            logging.debug('Added ' + item)
        else:
            logging.debug('Skipped ' + item)
            pass
    while True:
        print('\n')
        try:
            for pdf in pdf_docs:
                print(str(pdf_docs.index(pdf)) + ' ' + pdf)
            print('\nType number of the document you want to print.\n')
            pdf_number = int(input('Freemium/Freemium plus 1 volume document: '))
            print('Selected document: ' + pdf_docs[pdf_number])
            break
        except ValueError:
            print('Only numbers are acceptable. Please, try again.')
        except IndexError:
            print('No document with this number. Please, try again.')

    pp_payment = 'Paypal'
    cc_payment = 'Credit Card'
    payment_list = [pp_payment, cc_payment]
    while True:
        try:
            print('\n')
            for payment in payment_list:
                print(str(payment_list.index(payment)) + ' ' + payment)
            payment_num = str(input('\nEnter payment method number: '))
            print('Selected payment method: ' + payment_list[int(payment_num)])
            break
        except IndexError:
            print('Invalid input. Please, try again.')
        except ValueError:
            print('Only numbers are acceptable. Please, try again.')

    # Create instance and go to landing page:

    browser = webdriver.Chrome('C:\ChromeDriver\chromedriver.exe')
    browser.get('https://staging.pluspeter.com')
    browser.maximize_window()
    WebDriverWait(browser, 10).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    logging.debug('Login page rendered.')

    # Go to login page and sign in:

    browser.find_element_by_link_text('Login').click()
    logging.debug('Clicked on login button.')
    browser.find_element_by_id('id_username').send_keys(user_login)
    logging.debug('Typed login.')
    browser.find_element_by_id('id_password').send_keys(user_password)
    logging.debug('Typed password.')
    browser.find_element_by_id('id_password').submit()
    logging.debug('Submit login form.')

    # Go to products page and open freemium volume:

    WebDriverWait(browser, 10).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    logging.debug('Profile page rendered.')
    WebDriverWait(browser, 5).until(ec.element_to_be_clickable((By.LINK_TEXT, 'Jetzt Drucken'))).click()
    logging.debug('Clicked on "Print now" button.')
    WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    logging.debug('Product selection page rendered.')
    WebDriverWait(browser, 12).until(ec.element_to_be_clickable((By.XPATH,
                                                                 "//div[@class='button-bottom-holder']"
                                                                 "/button[@id='print_freemium']"))).click()
    logging.debug('Clicked on "Print now" button for Freemium.')

    # Add document:

    WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    logging.debug('Wizard rendered.')
    try:
        browser.find_element_by_id('omnibox-file-input').send_keys(test_docs_dir + pdf_docs[pdf_number])
        logging.debug('Added document ' + pdf_docs[pdf_number])
        WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
        logging.debug('Wizard page rendered.')
        browser.find_element_by_xpath('//*[@id="wizard"]/div[2]/div[1]/div/div[4]/input').send_keys('Free/mium by-b0t!')
        logging.debug('Typed volume title.')
        browser.find_element_by_xpath('//*[@id="wizard"]').click()
        logging.debug('Clicked on wizard.')
        WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
        logging.debug('Wizard page rendered.')
    except NoSuchElementException:
        logging.debug('Wizard page did not open.')
        browser.find_element_by_id('print_freemium_plus').click()
        logging.debug('Clicking on "Print now" button for freemium plus.')
        WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
        logging.debug('Wizard page rendered.')
        browser.find_element_by_name('file-input').send_keys(test_docs_dir + pdf_docs[pdf_number])
        logging.debug('Uploaded document ' + pdf_docs[pdf_number])

    if payment_list[int(payment_num)] == cc_payment:
        try:
            # Select CC payment method, mark checkboxes and open stripe:

            WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH,
                                                                                 "//div[@class='loading_box']")))
            logging.debug('Wizard page rendered.')
            WebDriverWait(browser, 2).until(ec.presence_of_element_located((By.ID, 'paypal3'))).click()
            logging.debug('Clicked on "Credit card" radiobutton')
            browser.find_element_by_xpath("//label[@for='checkbox-1']").click()
            logging.debug('Marked first checkbox.')
            browser.find_element_by_xpath("//label[@for='checkbox-2']").click()
            logging.debug('Marked second checkbox.')
            browser.find_element_by_id('paypalPayment').click()
            logging.debug('Clicked on "Print now" button.')

            # Switch to stripe and finish checkout:

            WebDriverWait(browser, 15).until(ec.frame_to_be_available_and_switch_to_it('stripe_checkout_app'))
            logging.debug('Switched to stripe overlay.')
            browser.find_element_by_xpath("///input[@placeholder='Card number']]").send_keys('4242 4242 4242 4242')
            logging.debug('Typed card number (Stripe).')
            browser.find_element_by_xpath("//input[@placeholder='MM / YY']").send_keys('0123')
            logging.debug('Typed month/year (Stripe).')
            browser.find_element_by_xpath("//input[@placeholder='CVC']").send_keys('456')
            logging.debug('Typed CVC (Stripe)')
            browser.find_element_by_xpath("//input[@placeholder='ZIP Code']").send_keys('7890')
            logging.debug('Typed zip code (Stripe).')
            browser.find_element_by_xpath("//button[@type='submit']").click()
            logging.debug('Submit Stripe.')
            browser.switch_to.default_content()
            logging.debug('Switched to default content.')

        except TimeoutException:
            logging.debug('Product is detected to be free.')

        # Mark checkboxes and finish checkout:

        try:
            WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH,
                                                                                 "//div[@class='loading_box']")))
            logging.debug('Wizard page rendered.')
            browser.find_element_by_xpath("//label[@for='checkbox-1']").click()
            logging.debug('Marked first checkbox.')
            browser.find_element_by_xpath("//label[@for='checkbox-2']").click()
            logging.debug('Marked second checkbox.')
            browser.find_element_by_id('paypalPayment').click()
            logging.debug('Clicked on "Print now" button.')
        except WebDriverException:
            pass
    else:
        try:
            # Mark checkboxes and go to Paypal:

            WebDriverWait(browser, 12).until(
                ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
            logging.debug('Wizard page opened.')
            WebDriverWait(browser, 2).until(ec.presence_of_element_located((By.ID, 'paypal3'))).click()
            logging.debug('Clicked on Credit card radiobutton.')
            browser.find_element_by_id('paypal1').click()
            logging.debug('Clicked on Paypal radiobutton.')
            browser.find_element_by_xpath("//label[@for='checkbox-1']").click()
            logging.debug('Marked first checkbox.')
            browser.find_element_by_xpath("//label[@for='checkbox-2']").click()
            logging.debug('Marked second checkbox.')
            browser.find_element_by_id('paypalPayment').click()
            logging.debug('Clicked on "Print now" button.')
            WebDriverWait(browser, 10).until(
                ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
            logging.debug('Wizard page rendered.')
            WebDriverWait(browser, 10).until(ec.url_contains('www.sandbox.paypal.com'))
            logging.debug('URL changed to "www.sandbox.paypal.com"')

            # PayPal checkout:

            WebDriverWait(browser, 25).until(ec.invisibility_of_element_located((By.ID, 'preloaderSpinner')))
            logging.debug('Paypal page rendered.')
            try:
                browser.find_element_by_xpath("//a[@class='btn full ng-binding']").click()
                logging.debug('Clicked on login button (Paypal).')
            except NoSuchElementException:
                logging.debug('Login button was not located.')
            WebDriverWait(browser, 25).until(ec.invisibility_of_element_located((By.ID, 'preloaderSpinner')))
            logging.debug('Paypal page rendered.')
            WebDriverWait(browser, 10).until(ec.visibility_of_element_located((By.NAME,
                                                                               'login_email'))).send_keys(paypal_login)
            logging.debug('Typed Paypal login.')
            WebDriverWait(browser, 10).until(ec.visibility_of_element_located((By.ID, 'btnNext'))).click()
            logging.debug('Clicked on "Next" button.')
            WebDriverWait(browser, 25).until(ec.invisibility_of_element_located((By.ID, 'preloaderSpinner')))
            logging.debug('Paypal page rendred.')
            WebDriverWait(browser, 10).until(ec.visibility_of_element_located((By.ID,
                                                                               'password'))).send_keys(paypal_password)
            logging.debug('Typed Paypal password.')
            browser.find_element_by_id('btnLogin').click()
            logging.debug('Clicked on Login button.')
            WebDriverWait(browser, 30).until(ec.invisibility_of_element_located((By.ID, 'preloaderSpinner')))
            logging.debug('Paypal rendered.')
            browser.find_element_by_id('confirmButtonTop').click()
            logging.debug('Clicked on "Continue" button.')
            WebDriverWait(browser, 25).until(ec.invisibility_of_element_located((By.ID, 'preloaderSpinner')))
            logging.debug('Paypal page rendered.')

        except TimeoutException:
            logging.debug('Product is detected to be free.')

            # Mark checkboxes and finish checkout:

        try:
            WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH,
                                                                                 "//div[@class='loading_box']")))
            logging.debug('Wizard page rendered.')
            browser.find_element_by_xpath("//label[@for='checkbox-1']").click()
            logging.debug('Marked first checkbox.')
            browser.find_element_by_xpath("//label[@for='checkbox-2']").click()
            logging.debug('Marked second checkbox.')
            browser.find_element_by_id('paypalPayment').click()
            logging.debug('Clicked on "Print now" button.')
        except WebDriverException:
            pass

    # Go to profile:

    WebDriverWait(browser, 20).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    logging.debug('Successful upload page rendered.')
    WebDriverWait(browser, 20).until(ec.presence_of_element_located((By.XPATH, "//a[@class='closer']"))).click()
    logging.debug('Clicked on "Close" button.')

    # Logout and quit:

    WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    logging.debug('Profile page rendered.')
    WebDriverWait(browser, 7).until(ec.presence_of_element_located((By.LINK_TEXT, "Abmelden"))).click()
    logging.debug('Clicked on Logout button.')
    WebDriverWait(browser, 12).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
    logging.debug('Landing page rendered.')
    browser.quit()


if __name__ == "__main__":
    while True:
        auto_mode = 'Auto'
        manual_mode = 'Manual'
        creds_list = [auto_mode, manual_mode]
        while True:
            try:
                print('\n')
                for mode in creds_list:
                    print(str(creds_list.index(mode)) + ' ' + mode)
                mode_num = str(input('\nEnter mode number (Auto will let bot select credentials; Manual will let you'
                                     ' enter credentials): '))
                print('\nSelected mode: ' + creds_list[int(mode_num)] + '\n')
                break
            except IndexError:
                print('Invalid input. Please, try again.')
            except ValueError:
                print('Only numbers are acceptable. Please, try again.')
        if creds_list[int(mode_num)] == auto_mode:
            country_list = ['Germany', 'Austria', 'Switzerland']
            while True:
                try:
                    for country in country_list:
                        print(str(country_list.index(country)) + ' ' + country)
                    user_country = str(input('\nEnter user\'s country number: '))
                    print('Selected country ' + country_list[int(user_country)])
                    break
                except IndexError:
                    print('Invalid input. Please, try again.')
                except ValueError:
                    print('Only numbers are acceptable. Please, try again.')
            print('Program will select creds automatically.')
            os.chdir(docs_dir)
            login_list = []
            password_list = []
            wb = openpyxl.load_workbook('credentials.xlsx')
            wb.save('credentials - Copy.xlsx')
            wb = openpyxl.load_workbook('credentials - Copy.xlsx')
            if country_list[int(user_country)] == 'Germany':
                sheet = wb['German users']
                logging.debug('Selected German users worksheet.')
            elif country_list[int(user_country)] == 'Austria':
                sheet = wb['Austrian users']
                logging.debug('Selected Austrian users worksheet.')
            else:
                sheet = wb['Swiss users']
                logging.debug('Selected Swiss users worksheet.')
            if sheet.max_row <= 2:
                logging.debug('Located equal to or less then two rows in worksheet.')
                login_list.append(sheet.cell(2, 1).value)
                logging.debug('Appending login cell value: ' + sheet.cell(2, 1).value)
                password_list.append(sheet.cell(2, 2).value)
                logging.debug('Appending password cell value: ' + sheet.cell(2, 2).value)
                login = login_list[0]
                logging.debug('Selected login: ' + login_list[0])
                password = password_list[0]
                logging.debug('Selected password: ' + password_list[0])
            else:
                for row in range(2, sheet.max_row + 1):
                    login_list.append(sheet.cell(row, 1).value)
                    password_list.append(sheet.cell(row, 2).value)
                logging.debug('Login list length is: ' + str(len(login_list)))
                logging.debug('Password list length is: ' + str(len(password_list)))
                creds_num = random.randint(0, len(login_list) - 1)
                logging.debug('Selected credentials number ' + str(creds_num))
                login = login_list[creds_num]
                logging.debug('Selected login: ' + login_list[creds_num])
                password = password_list[creds_num]
                logging.debug('Selected password: ' + password_list[creds_num])
            break
        elif creds_list[int(mode_num)] == manual_mode:
            print('Please, enter requested credentials.')
            login = input('Login: ')
            password = input('Password: ')
            break
        else:
            print('Invalid input. Please, try again.')

    init_checkout(login, password)
