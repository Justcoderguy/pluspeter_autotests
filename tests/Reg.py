#! python3
# Reg - registration flow program


import random
import string
import openpyxl
import os
import logging
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as ec

logging.basicConfig(level=logging.DEBUG, filename='Reg.log', filemode='w',
                    format='%(asctime)s - %(levelname)s - %(message)s', datefmt='%d/%m/%Y %I:%M:%S')

logging.disable(logging.DEBUG)  # comment out to enable logs

logging.debug('Start of the program.')
de = ('plusdomain.de', 'plusdomain.com', 'anydomain.de', 'deutschdomain.de')
aus = ('ausdomain.de',)
chz = ('swisdomain.de',)

anrede = ('f', 'm')
profile_title = random.choice(anrede)
names = ('Ada', 'Adalbert', 'Baltasar', 'Bathilda', 'Charles', 'Claudette', 'Dedrick', 'Delmira',
         'Eburhardt', 'Elke', 'Franziska', 'Friedrick', 'Gabriele', 'Garrick', 'Halfrid', 'Hamlett',
         'Immanuel', 'Idette', 'Jakob', 'Jolan')
profile_name = random.choice(names)
surnames = ('Müller', 'Schmidt', 'Schneider', 'Fischer', 'Weber', 'Mayer', 'Becker', 'Schulz', 'Hoffmann',
            'Schäfer', 'Koch', 'Bauer', 'Richter', 'Klein', 'Wolf', 'Schröder', 'Neumann', 'Schwarz',
            'Zimmermann', 'Braun', 'Krüger', 'Groß', 'Jäger')
profile_surname = random.choice(surnames)
streets = ('Ackerstraße', 'Bernauer Straße', 'Fasanerieallee', 'Straße des 17. Juni', 'Gollanczstraße',
           'Oranienburger Straße', 'Tangentiale Verbindung Ost', 'Straße der Pariser Kommune',
           'Tangentiale Verbindung Ost', 'Voßstraße', 'Rosa-Luxemburg-Straße')
profile_street = random.choice(streets)
cities = ('Berlin', 'München', 'Düsseldorf', 'Köln', 'Nürnberg', 'Münster', 'Mülheim an der Ruhr', 'Recklinghausen')
profile_city = random.choice(cities)
ger_phone = ('491123456789', '4911234567890', '+491123456789',
             '+4911234567890', '00491123456789', '004911234567890')
profile_ger_phone = random.choice(ger_phone)
ger_domains = ('gmx.org', 'web.de', 't-online.net', 'outlook.com.de', 'gmail.com')
private_email = ''.join(random.choice(string.hexdigits) for char in range(20)) \
                + '.student@' + random.choice(ger_domains)
logging.debug('Private email generated: %s' % private_email)
aus_ch_phone = '+' + ''.join(random.choice(string.digits) for p_num in range(random.randint(1, 19)))
logging.debug('Phone number (Austria/Switzerland) generated: %s' % aus_ch_phone)
house_num = ''.join(random.choice(string.digits) for h_num in range(random.randint(1, 12)))
logging.debug('House number generated: %s' % house_num)
extra_num = ''.join(random.choice(string.hexdigits) for e_num in range(random.randint(1, 60)))
logging.debug('Additional address number generated: %s' % extra_num)
zip_ger = ''.join(random.choice(string.digits) for ger_number in range(5))
logging.debug('Zip code (Germany) generated: %s' % zip_ger)
zip_aus_ch = ''.join(random.choice(string.digits) for aus_ch_number in range(4))
logging.debug('Zip code (Austria/Switzerland) generated: %s' % zip_aus_ch)
birth_dates = str(random.randint(1, 28)) + '.' + str(random.randint(1, 12)) + '.' + str(random.randint(1973, 2002))
logging.debug('Birth date generated: %s' % birth_dates)
semesters = random.randint(1, 12)
docs_dir = 'C:\\Users\\pzakharevich\\Desktop\\Education\\PlusPeter\\docs'
logging.info('Documents directory defined: %s' % docs_dir)
country_1 = 'Germany'
country_2 = 'Austria'
country_3 = 'Switzerland'
country_list = [country_1, country_2, country_3]
while True:
    print('\n')
    for item in country_list:
        print(str(country_list.index(item)) + ' ' + item)
    while True:
        try:
            country_num = str(input('\nEnter number of the country: '))
            print('Selected country: ' + country_list[int(country_num)])
            break
        except IndexError:
            print('Invalid input. Please, try again.')
        except ValueError:
            print('Only numbers are acceptable. Please, try again.')
    if country_list[int(country_num)] == country_1:
        domains = de
        break
    elif country_list[int(country_num)] == country_2:
        domains = aus
        break
    elif country_list[int(country_num)] == country_3:
        domains = chz
        break
    else:
        print('Invalid input. Please, try again.')
login = ''.join(random.choice(string.hexdigits) for char in range(20)) + '.student@' + random.choice(domains)
logging.debug('Login generated: %s' % login)
password = ''.join(random.choice(string.hexdigits + string.punctuation) for char in range(8))
logging.debug('Password generated: %s' % password)
print('\nLogin: ' + login + '\nPassword: ' + password)

# Create instance and go to landing page:

browser = webdriver.Chrome('C:\ChromeDriver\chromedriver.exe')
browser.get('https://staging.pluspeter.com')
browser.maximize_window()

# Go to registration page and fill form:

WebDriverWait(browser, 10).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
logging.debug('Landing page rendered.')
browser.find_element_by_xpath('/html/body/div[1]/div/header/div/div[1]/div/a[1]').click()
logging.debug('Opened registration page.')
WebDriverWait(browser, 10).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
logging.debug('Registration page rendered.')
browser.find_element_by_xpath('//*[@id="id_email"]').send_keys(login)
logging.debug('Typed login.')
browser.find_element_by_xpath('//*[@id="id_password"]').send_keys(password)
logging.debug('Typed password.')
browser.find_element_by_xpath('//*[@id="id_password_repeat"]').send_keys(password)
logging.debug('Typed password in confirmation field.')
browser.find_element_by_xpath('/html/body/div[1]/div/main/div/div[2]/form/div[1]/label[1]/span').click()
logging.debug('Marked first checkbox.')
browser.find_element_by_xpath('/html/body/div[1]/div/main/div/div[2]/form/div[1]/label[2]/span').click()
logging.debug('Marked second checkbox.')

# Finish registration:

browser.find_element_by_xpath('/html/body/div[1]/div/main/div/div[2]/form/div[2]/button').click()
logging.debug('Submit registration form.')
input('\nPlease, follow the activation link you should receive from email. '
      'After that, switch to this window and press "Enter".')

# log in and fill profile:

browser.get('https://staging.pluspeter.com/students/login/')
WebDriverWait(browser, 10).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
browser.find_element_by_id('id_username').send_keys(login)
browser.find_element_by_id('id_password').send_keys(password)
browser.find_element_by_xpath('//*[@type="submit"]').click()
WebDriverWait(browser, 20).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
WebDriverWait(browser, 10).until(ec.element_to_be_clickable((By.ID, 'title_input'))).click()
browser.find_element_by_xpath("//div[@value='" + profile_title + "']").click()
logging.info('Selected title: ' + profile_title)
browser.find_element_by_id('datum').send_keys(birth_dates)
logging.info('Typed birth date: ' + birth_dates)
browser.find_element_by_id('vorname').send_keys(profile_name)
logging.info('Typed name: ' + profile_name)
browser.find_element_by_id('name').send_keys(profile_surname)
logging.info('Typed surname: ' + profile_surname)
browser.find_element_by_id('privateemail').send_keys(private_email)
logging.info('Typed private email: ' + private_email)
if domains == de:
    browser.find_element_by_id('Handynummer').send_keys(profile_ger_phone)
    logging.info('Typed phone number: ' + profile_ger_phone)
else:
    browser.find_element_by_id('Handynummer').send_keys(aus_ch_phone)
    logging.info('Typed phone number: ' + aus_ch_phone)
browser.find_element_by_id('Adresse').send_keys(profile_street)
logging.info('Typed address: ' + profile_street)
browser.find_element_by_id('Hausnummer').send_keys(house_num)
logging.info('Typed house number: ' + house_num)
browser.find_element_by_id('Zusatz').send_keys(extra_num)
logging.info('Typed additional address: ' + str(extra_num))
browser.find_element_by_id('Stadt').send_keys(profile_city)
logging.info('Typed city:' + profile_city)
if domains == de:
    browser.find_element_by_id('zip_code').send_keys(zip_ger)
    logging.info('Typed zip code: ' + zip_ger)
else:
    browser.find_element_by_id('zip_code').send_keys(zip_aus_ch)
    logging.info('Typed zip code: ' + zip_aus_ch)
browser.find_element_by_id('education_input').click()
browser.find_element_by_xpath("//div[@value=68]").click()
logging.debug('Selected education.')
browser.find_element_by_id('study_input').click()
browser.find_element_by_xpath("//div[@value=2]").click()
logging.debug('Selected study.')
browser.find_element_by_id('semester_input').click()
browser.find_element_by_xpath("//div[@value=" + str(semesters) + "]").click()
logging.info('Selected semester:' + str(semesters))
browser.find_element_by_id("profile_update_btn").click()
logging.debug('Saved popup.')
WebDriverWait(browser, 20).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
logging.debug('Profile page rendered.')
WebDriverWait(browser, 10).until(ec.element_to_be_clickable((By.ID, 'profile_thanks_btn'))).click()
logging.debug('Closed thanks popup.')
WebDriverWait(browser, 20).until(ec.invisibility_of_element_located((By.XPATH, "//div[@class='loading_box']")))
logging.debug('Profile page rendered.')
browser.find_element_by_xpath("//a[@class='btn btn-stroke login']").click()
logging.debug('Logout.')
browser.quit()

# Save credentials into excel file:

os.chdir(docs_dir)
wb = openpyxl.load_workbook('credentials.xlsx')
logging.debug('Loaded workbook.')
if domains == de:
    sheet = wb['German users']
    logging.debug('Credentials defined as German.')
elif domains == aus:
    sheet = wb['Austrian users']
    logging.debug('Credentials defined as Austrian.')
else:
    sheet = wb['Swiss users']
    logging.debug('Credentials defined as Swiss.')
sheet.cell(sheet.max_row + 1, 1).value = login
sheet.cell(sheet.max_row, 2).value = password
wb.save('credentials.xlsx')
logging.debug('Credentials are saved successfully in ' + sheet.title)
